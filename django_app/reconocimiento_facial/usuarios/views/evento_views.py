"""
-----------------------------------------------------------------------------
Archivo: evento_views.py
Descripcion: Controlador CRUD para gestion de eventos academicos.
             Permite crear, editar, eliminar y listar eventos.
             Actualiza automaticamente el estado (pendiente/activo/finalizado)
             segun fecha y hora. Exporta planillas de asistencia a Excel.
Fecha de creacion: 01 de Noviembre 2025
Fecha de modificacion: 20 de Diciembre 2025
Autores:
    Roberto Leal
    William Tapia
-----------------------------------------------------------------------------
"""

from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from ..services.firebase_service import firebase_service
from ..decorators import admin_required
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill



@admin_required
def listar_eventos(request):
    """Lista todos los eventos ordenados por fecha."""
    try:
        # Actualizar estados automáticamente antes de listar
        firebase_service.actualizar_estados_eventos()
        
        # Obtener eventos desde Firebase
        eventos = firebase_service.listar_eventos()
        print(f"🔍 DEBUG: Total eventos recibidos de Firebase: {len(eventos)}")
        if eventos:
            print(f"🔍 DEBUG: Primer evento: {eventos[0]}")
        
        # Obtener parámetro de búsqueda
        search_query = request.GET.get('search', '').strip()
        
        # Aplicar búsqueda por nombre de evento si existe
        if search_query:
            search_lower = search_query.lower()
            eventos = [e for e in eventos if search_lower in e.get('nombre', '').lower()]
        
        # Filtros básicos en memoria (Firebase tiene limitaciones de query complejos)
        year_filter = request.GET.get('year')
        month_filter = request.GET.get('month')
        semestre_filter = request.GET.get('semestre')
        
        # Obtener todos los eventos sin filtro primero para estadísticas
        todos_eventos = list(eventos)
        
        # Aplicar filtros combinados correctamente
        if year_filter and month_filter:
            # Filtro combinado: año Y mes
            try:
                month_num = int(month_filter)
                eventos = [e for e in eventos if e.get('fecha', '').startswith(year_filter) and int(e.get('fecha', '').split('-')[1]) == month_num]
            except (ValueError, IndexError):
                pass
        elif year_filter:
            # Solo filtro de año
            eventos = [e for e in eventos if e.get('fecha', '').startswith(year_filter)]
        elif month_filter:
            # Solo filtro de mes (todos los años)
            try:
                month_num = int(month_filter)
                eventos = [e for e in eventos if int(e.get('fecha', '').split('-')[1]) == month_num]
            except (ValueError, IndexError):
                pass
            
        if semestre_filter:
            def get_month(fecha_str):
                try:
                    return int(fecha_str.split('-')[1])
                except:
                    return 0
                
            if semestre_filter == 'otono':  # Marzo (3) - Julio (7)
                eventos = [e for e in eventos if 3 <= get_month(e.get('fecha', '')) <= 7]
            elif semestre_filter == 'primavera':  # Agosto (8) - Diciembre (12)
                eventos = [e for e in eventos if 8 <= get_month(e.get('fecha', '')) <= 12]

        # Obtener años disponibles con contadores
        years = {}
        for e in todos_eventos:
            try:
                fecha = e.get('fecha', '')
                if fecha:
                    year = fecha.split('-')[0]
                    years[year] = years.get(year, 0) + 1
            except:
                pass
        available_years = [{'year': y, 'count': c} for y, c in sorted(years.items(), reverse=True)]
        
        # Meses con nombres en español y contadores
        meses_base = [
            {'num': 1, 'nombre': 'Enero'},
            {'num': 2, 'nombre': 'Febrero'},
            {'num': 3, 'nombre': 'Marzo'},
            {'num': 4, 'nombre': 'Abril'},
            {'num': 5, 'nombre': 'Mayo'},
            {'num': 6, 'nombre': 'Junio'},
            {'num': 7, 'nombre': 'Julio'},
            {'num': 8, 'nombre': 'Agosto'},
            {'num': 9, 'nombre': 'Septiembre'},
            {'num': 10, 'nombre': 'Octubre'},
            {'num': 11, 'nombre': 'Noviembre'},
            {'num': 12, 'nombre': 'Diciembre'},
        ]
        
        # Calcular contadores por mes (considerando año si está seleccionado)
        for mes in meses_base:
            count = 0
            for e in todos_eventos:
                try:
                    fecha = e.get('fecha', '')
                    if fecha:
                        parts = fecha.split('-')
                        event_month = int(parts[1])
                        event_year = parts[0]
                        
                        # Si hay año seleccionado, contar solo eventos de ese año
                        if year_filter:
                            if event_year == year_filter and event_month == mes['num']:
                                count += 1
                        else:
                            # Sin año, contar todos los eventos de ese mes
                            if event_month == mes['num']:
                                count += 1
                except:
                    pass
            mes['count'] = count

        context = {
            'lista_eventos_completa': eventos,
            'available_years': available_years,
            'available_months': meses_base,
            'year_selected': int(year_filter) if year_filter else None,
            'month_selected': int(month_filter) if month_filter else None,
            'semestre_selected': semestre_filter,
            'search_query': search_query
        }
        print(f"🔍 DEBUG: Context eventos length: {len(context['lista_eventos_completa'])}")
        print(f"🔍 DEBUG: Context keys: {context.keys()}")
        return render(request, 'eventos_grid.html', context)
        
    except Exception as e:
        import traceback
        print(f"❌ ERROR CRÍTICO EN VISTA LISTAR_EVENTOS: {e}")
        traceback.print_exc()
        messages.error(request, f"Error listando eventos: {e}")
        return render(request, 'eventos_grid.html', {'lista_eventos_completa': []})


@admin_required
def crear_evento(request):
    """Crea un nuevo evento."""
    if request.method == 'POST':
        try:
            firebase_service.crear_evento(
                nombre=request.POST.get('nom_evento'),
                descripcion=request.POST.get('descripcion'),
                fecha=request.POST.get('fecha'),
                hora_inicio=request.POST.get('hora_inicio'),
                hora_fin=request.POST.get('hora_fin'),
                relator=request.POST.get('relator'),
                ubicacion=request.POST.get('ubicacion')
            )
            messages.success(request, "Evento creado correctamente")
            return redirect('listar_eventos')
        except Exception as e:
            messages.error(request, f"Error creando evento: {e}")
            
    return render(request, 'crear_evento.html')


@admin_required
def editar_evento(request, evento_id):
    """Edita un evento existente."""
    if request.method == 'POST':
        try:
            firebase_service.actualizar_evento(
                evento_id,
                nombre=request.POST.get('nom_evento'),
                descripcion=request.POST.get('descripcion'),
                fecha=request.POST.get('fecha'),
                hora_inicio=request.POST.get('hora_inicio'),
                hora_fin=request.POST.get('hora_fin'),
                relator=request.POST.get('relator'),
                ubicacion=request.POST.get('ubicacion')
            )
            messages.success(request, "Evento actualizado correctamente")
            return redirect('listar_eventos')
        except Exception as e:
            messages.error(request, f"Error actualizando evento: {e}")

    # GET: Obtener datos para el formulario
    evento = firebase_service.obtener_evento(evento_id)
    if not evento:
        messages.error(request, "Evento no encontrado")
        return redirect('listar_eventos')
        
    return render(request, 'editar_evento.html', {'evento': evento})


@admin_required
def eliminar_evento(request, evento_id):
    """Elimina un evento."""
    if request.method == 'POST':
        try:
            firebase_service.eliminar_evento(evento_id)
            messages.success(request, "Evento eliminado correctamente")
            return redirect('listar_eventos')
        except Exception as e:
            messages.error(request, f"Error eliminando evento: {e}")
            return redirect('listar_eventos')
    
    # GET: Pedir confirmación
    evento = firebase_service.obtener_evento(evento_id)
    if not evento:
        messages.error(request, "Evento no encontrado")
        return redirect('listar_eventos')
    
    return render(request, 'confirmar_eliminar_evento.html', {'evento': evento})


@admin_required
def descargar_planilla_evento(request, evento_id):
    """Descarga un archivo Excel con la lista de asistentes al evento."""
    try:
        # Obtener datos del evento
        evento = firebase_service.obtener_evento(evento_id)
        if not evento:
            messages.error(request, "Evento no encontrado")
            return redirect('listar_eventos')
        
        # Obtener asistencias del evento
        asistencias = firebase_service.listar_asistencias(id_evento=evento_id)
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistentes"
        
        # Configurar título
        nombre_evento = evento.get('nombre', 'Evento')
        ws.merge_cells('A1:G1')
        titulo_cell = ws['A1']
        titulo_cell.value = f"Lista de Asistentes - {nombre_evento}"
        titulo_cell.font = Font(size=14, bold=True, color="FFFFFF")
        titulo_cell.fill = PatternFill(start_color="B82020", end_color="B82020", fill_type="solid")
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Configurar encabezados
        encabezados = ['N°', 'RUT', 'Nombre', 'Carrera', 'Jornada', 'Fecha y Hora', 'Método']
        ws.append(encabezados)
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="DE4949", end_color="DE4949", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, 8):
            cell = ws.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
        
        # Agregar datos de asistencias
        for idx, asistencia in enumerate(asistencias, start=1):
            # Formatear fecha y hora
            fecha_hora = asistencia.get('fecha_hora', '')
            if 'T' in fecha_hora:
                fecha_hora = fecha_hora.replace('T', ' ')[:19]
            
            # Formatear método
            metodo = asistencia.get('metodo', 'manual')
            metodo_texto = 'Biométrico' if metodo == 'biometrico' else 'Manual'
            
            # Formatear jornada
            jornada = asistencia.get('jornada_usuario', '')
            jornada_texto = 'Diurna' if jornada == 'D' else 'Vespertina' if jornada == 'V' else jornada
            
            fila = [
                idx,
                asistencia.get('rut_usuario', ''),
                asistencia.get('nombre_usuario', ''),
                asistencia.get('carrera_usuario', ''),
                jornada_texto,
                fecha_hora,
                metodo_texto
            ]
            ws.append(fila)
            
            # Aplicar bordes a las celdas de datos
            for col in range(1, 8):
                cell = ws.cell(row=idx+2, column=col)
                cell.border = border_style
                if col == 1:  # N° centrado
                    cell.alignment = Alignment(horizontal='center')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Limpiar nombre del evento para usar en el archivo
        nombre_archivo = nombre_evento.replace(' ', '_').replace('/', '-')
        response['Content-Disposition'] = f'attachment; filename="Asistentes_{nombre_archivo}.xlsx"'
        
        # Guardar el libro en la respuesta
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f"Error generando planilla: {e}")
        return redirect('listar_eventos')

